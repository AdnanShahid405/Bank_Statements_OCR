"""
excel_writer.py

Writes extracted transactions to a structured Excel file.
- One sheet per PDF
- One row per transaction
- Red cell fill for any field where Azure confidence score < 0.85
- Validation summary appended at the bottom of each sheet
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from extractor import ExtractionResult, Transaction
from validator import ValidationResult

# ── Constants ──────────────────────────────────────────────────────────────────

CONFIDENCE_THRESHOLD = 0.85

# Colors
RED_FILL    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
ALT_FILL    = PatternFill(start_color="EEF3FB", end_color="EEF3FB", fill_type="solid")
SUMMARY_FILL= PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WARN_FILL   = PatternFill(start_color="FFE599", end_color="FFE599", fill_type="solid")

WHITE_FONT  = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
BOLD_FONT   = Font(name="Calibri", bold=True, size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
RED_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COLUMNS = ["Date", "Description", "Debit ($)", "Credit ($)", "Balance ($)"]
COL_WIDTHS = [16, 40, 14, 14, 16]

# Maps column index (0-based) → Transaction confidence attribute
CONF_MAP = {
    0: "date_conf",
    1: "description_conf",
    2: "debit_conf",
    3: "credit_conf",
    4: "balance_conf",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _apply_border(cell):
    cell.border = THIN_BORDER


def _fmt(val, is_amount=False):
    if val is None:
        return "—"
    if is_amount:
        return val  
    return val


def _low_conf_fields(txn: Transaction) -> list:
    """Returns list of field names with confidence below threshold."""
    low = []
    labels = {
        "date_conf": "Date",
        "description_conf": "Description",
        "debit_conf": "Debit",
        "credit_conf": "Credit",
        "balance_conf": "Balance",
    }
    for attr, label in labels.items():
        val = getattr(txn, attr)
        if val is not None and val < CONFIDENCE_THRESHOLD:
            low.append(f"{label} ({val:.0%})")
    return low


# ── Sheet Builder ──────────────────────────────────────────────────────────────

def _write_sheet(ws, result: ExtractionResult, validation: ValidationResult):
    is_azure = result.extraction_method == "azure"
    total_cols = len(COLUMNS)

    # ── Sheet title ───────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
    title_cell = ws["A1"]
    title_cell.value = f"Statement: {result.pdf_name}  |  Method: {result.extraction_method.upper()}"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="003366")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # ── Header row ────────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(zip(COLUMNS, COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        _apply_border(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[2].height = 20

    # ── Transaction rows ──────────────────────────────────────────────────────
    for row_idx, txn in enumerate(result.transactions, start=3):
        is_alt = (row_idx % 2 == 0)

        row_data = [
            txn.date,
            txn.description,
            txn.debit,
            txn.credit,
            txn.balance,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = NORMAL_FONT
            cell.alignment = Alignment(
                horizontal="right" if col_idx in (3, 4, 5) else "left",
                vertical="center"
            )
            _apply_border(cell)

            if col_idx in (3, 4, 5) and isinstance(value, float):
                cell.number_format = '#,##0.00'

            # Red fill for low-confidence fields (Azure only)
            conf_attr = CONF_MAP.get(col_idx - 1)
            if conf_attr and is_azure:
                conf_val = getattr(txn, conf_attr)
                if conf_val is not None and conf_val < CONFIDENCE_THRESHOLD:
                    cell.fill = RED_FILL
                    cell.font = RED_FONT
                    continue

            if is_alt:
                cell.fill = ALT_FILL

        ws.row_dimensions[row_idx].height = 18

    # ── Validation Summary ────────────────────────────────────────────────────
    summary_start = len(result.transactions) + 4

    ws.cell(row=summary_start, column=1, value="BALANCE VALIDATION SUMMARY").font = BOLD_FONT

    summary_rows = [
        ("Opening Balance",  validation.opening_balance,  False),
        ("Total Credits",    validation.total_credits,    False),
        ("Total Debits",     validation.total_debits,     False),
        ("Expected Closing", validation.expected_closing, False),
        ("Actual Closing",   validation.actual_closing,   False),
        ("Status",           "✅ PASS" if validation.is_valid else "❌ FAIL", True),
    ]
    if not validation.is_valid and validation.variance is not None:
        summary_rows.append(("Variance", validation.variance, True))

    for i, (label, value, is_flag) in enumerate(summary_rows):
        r = summary_start + 1 + i
        label_cell = ws.cell(row=r, column=1, value=label)
        value_cell = ws.cell(row=r, column=2, value=value)

        label_cell.font = BOLD_FONT
        label_cell.fill = SUMMARY_FILL
        value_cell.fill = SUMMARY_FILL
        _apply_border(label_cell)
        _apply_border(value_cell)

        if isinstance(value, float):
            value_cell.number_format = '#,##0.00'

        if is_flag and not validation.is_valid:
            value_cell.fill = WARN_FILL
            value_cell.font = Font(name="Calibri", bold=True, color="CC0000", size=10)


# ── Public Entry Point ─────────────────────────────────────────────────────────

def write_excel(
    results: list,          
    validations: list,      
    output_path: str
):
    """
    Creates an Excel file with one sheet per PDF.
    results and validations must be in the same order.
    """
    wb = Workbook()
    wb.remove(wb.active)  

    for result, validation in zip(results, validations):
        sheet_name = result.pdf_name.replace(".pdf", "")[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, result, validation)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"\nExcel file saved: {output_path}")